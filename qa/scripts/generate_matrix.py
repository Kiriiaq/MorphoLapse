#!/usr/bin/env python3
"""Génère qa/matrice_tests.xlsx (118 tests + feuille Synthèse).

Source unique de vérité pour la matrice de qualification MorphoLapse v2.0.0.
Régénérer à chaque évolution de la liste de tests.

Usage :
    python qa/scripts/generate_matrix.py
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TEST_DIR = Path(__file__).resolve().parent.parent
OUT = TEST_DIR / "matrice_tests.xlsx"

# Catégories couleur (header)
CAT_FILL = {
    "IHM": "FFE5F0FB",          # bleu très clair
    "Paramètres": "FFE7F6E5",   # vert très clair
    "Entrées": "FFFFF4E0",      # ambre clair
    "Sorties": "FFEFE5FA",      # violet clair
    "Cas limites": "FFFFE0E0",  # rouge clair
    "Performance": "FFE0F2F4",  # cyan clair
    "Robustesse": "FFFFE0CC",   # orange clair
    "Régression": "FFE5E5E5",   # gris clair
}

HEADER_FILL = PatternFill("solid", fgColor="FF2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# (id, cat, req, fonctionnalité, description, prérequis, donnée entrée, résultat attendu, sévérité)
TESTS: list[tuple[str, str, str, str, str, str, str, str, str]] = [
    # ============ IHM (T-001..T-040) ============
    ("T-001", "IHM", "—", "Démarrage GUI", "Lancement python main.py : splash 5 étapes puis fenêtre principale visible (90% écran, centrée)", "—", "—", "Splash apparaît, fenêtre principale visible sans crash, stderr vide", "Bloquant"),
    ("T-002", "IHM", "REQ-001 / REQ-013", "Sélecteur dossier source", "Cliquer '...' à droite de 'Dossier source' puis sélectionner un dossier valide", "App démarrée", "qa/inputs/input_nominal", "Champ peuplé, previews Début/Fin affichées, stats '5 images | Réf: Auto | Sortie: -'", "Bloquant"),
    ("T-003", "IHM", "REQ-002", "Sélecteur image référence", "Cliquer '...' à droite de 'Image de référence' et choisir un PNG", "App démarrée", "qa/inputs/input_nominal/face_000.png", "Champ peuplé, stats reflète 'Réf: face_000.png'", "Majeur"),
    ("T-004", "IHM", "REQ-002", "Sélecteur dossier sortie", "Cliquer '...' à droite de 'Dossier de sortie' et choisir un dossier", "App démarrée", "qa/outputs_reels/manual_T004", "Champ peuplé, stats reflète 'Sortie: ✓'", "Majeur"),
    ("T-005", "IHM", "REQ-008", "StepIndicator Import - toggle", "Décocher puis recocher la case de l'étape 'Import des images'", "App démarrée", "—", "Statut bascule pending↔disabled visuellement (icône + couleur)", "Majeur"),
    ("T-006", "IHM", "REQ-008", "StepIndicator Import - statut running", "Lancer un workflow et observer l'indicateur Import pendant son exécution", "input_nominal sélectionné", "input_nominal", "Indicateur Import passe au statut 'running' (icône ◉, bordure bleue)", "Majeur"),
    ("T-007", "IHM", "REQ-008", "StepIndicator Import - mini progress", "Pendant le run, observer la mini barre 60×6 à droite de chaque step", "Workflow lancé", "input_nominal", "Mini barre se remplit progressivement", "Mineur"),
    ("T-008", "IHM", "REQ-008", "StepIndicator Align - toggle", "Idem T-005 sur 'Alignement des visages'", "App démarrée", "—", "Statut bascule correctement", "Majeur"),
    ("T-009", "IHM", "REQ-008", "StepIndicator Align - statut completed", "Workflow nominal, observer fin étape align", "input_nominal", "input_nominal", "Indicateur passe à 'completed' (✓ vert, progress 100)", "Majeur"),
    ("T-010", "IHM", "REQ-008", "StepIndicator Align - statut error", "Image sans visage à l'index 2, retry_detection=1, continue_on_error=False", "input_no_face", "input_no_face", "Indicateur passe à 'error' (✗ rouge)", "Majeur"),
    ("T-011", "IHM", "REQ-008", "StepIndicator Morph - toggle", "Idem sur 'Morphing facial'", "—", "—", "OK", "Majeur"),
    ("T-012", "IHM", "REQ-008", "StepIndicator Morph - running", "Idem T-006 sur Morph", "—", "input_nominal", "OK", "Majeur"),
    ("T-013", "IHM", "REQ-005 / REQ-015", "StepIndicator Morph - statut cancelled", "Lancer workflow et annuler pendant Morph (Échap ou clic Annuler)", "input_nominal (≥5 img)", "input_nominal", "Indicateur Morph passe à 'cancelled' (⏹ ambre)", "Bloquant"),
    ("T-014", "IHM", "REQ-008", "StepIndicator Export - toggle", "—", "—", "—", "OK", "Mineur"),
    ("T-015", "IHM", "REQ-008", "StepIndicator Export - completed", "Workflow complet jusqu'à export", "input_nominal", "input_nominal", "OK", "Majeur"),
    ("T-016", "IHM", "REQ-008", "StepIndicator Export - skipped", "Décocher Export puis lancer", "input_nominal", "input_nominal", "Indicateur passe à 'skipped' (⊘ gris)", "Mineur"),
    ("T-017", "IHM", "REQ-003", "Bouton Lancer disabled à vide", "Vider config.json puis démarrer l'app sans sélection", "Config réinitialisée", "—", "Bouton Lancer grisé/disabled", "Bloquant"),
    ("T-018", "IHM", "REQ-003", "Bouton Lancer s'active à sélection", "Avec champ vide, sélectionner un dossier source", "—", "qa/inputs/input_nominal", "Lancer passe à enabled instantanément (trace)", "Bloquant"),
    ("T-019", "IHM", "REQ-003", "Bouton Lancer se redésactive", "Avec dossier sélectionné, vider manuellement le champ Entry", "Dossier source sélectionné", "—", "Lancer redevient disabled", "Mineur"),
    ("T-020", "IHM", "REQ-003 / REQ-016", "Bouton Lancer disabled pendant run", "Démarrer un workflow", "input_nominal sélectionné", "input_nominal", "Lancer disabled pendant tout le run", "Majeur"),
    ("T-021", "IHM", "REQ-004", "Bouton Annuler disabled au démarrage", "Au lancement de l'app", "—", "—", "Annuler affiché en rouge mais disabled", "Majeur"),
    ("T-022", "IHM", "REQ-004", "Bouton Annuler enabled pendant run", "Démarrer un workflow", "input_nominal", "input_nominal", "Annuler enabled pendant le run", "Majeur"),
    ("T-023", "IHM", "REQ-004", "Bouton Annuler redisabled après run", "Attendre la fin du workflow", "input_nominal", "input_nominal", "Annuler redevient disabled", "Mineur"),
    ("T-024", "IHM", "REQ-006", "Tooltip Annuler mentionne Échap", "Hover ≥ 400 ms sur le bouton Annuler", "—", "—", "Tooltip affiche 'Annule le traitement en cours (Échap)'", "Mineur"),
    ("T-025", "IHM", "—", "QuickAction Ouvrir 📂", "Cliquer l'icône 📂 dans la toolbar centrale", "App démarrée", "—", "Ouvre filedialog de sélection dossier source", "Mineur"),
    ("T-026", "IHM", "REQ-011", "QuickAction Sauvegarder 💾", "Cliquer l'icône 💾 dans la toolbar centrale", "App démarrée", "—", "Sauvegarde immédiate (log 'Paramètres sauvegardés')", "Mineur"),
    ("T-027", "IHM", "REQ-002", "stats_label après sélection input", "Sélectionner input_nominal", "App démarrée", "input_nominal (5 img)", "Label '5 images | Réf: Auto | Sortie: -'", "Bloquant"),
    ("T-028", "IHM", "REQ-002", "stats_label après sélection référence", "Avec input_nominal sélectionné, choisir une référence", "input_nominal sélectionné", "face_000.png", "Label '5 images | Réf: face_000.png | Sortie: -'", "Bloquant"),
    ("T-029", "IHM", "REQ-002", "stats_label après sélection sortie", "Sélectionner un dossier de sortie quelconque", "input_nominal sélectionné", "manual_T029/", "Label '5 images | Réf: ... | Sortie: ✓'", "Majeur"),
    ("T-030", "IHM", "REQ-013", "Preview Début après sélection", "Sélectionner input_nominal", "App démarrée", "input_nominal", "Vignette 100×100 'face_000.png' visible sous 'Début'", "Majeur"),
    ("T-031", "IHM", "REQ-013", "Preview Fin après sélection", "Idem T-030", "App démarrée", "input_nominal", "Vignette 100×100 'face_004.png' visible sous 'Fin'", "Majeur"),
    ("T-032", "IHM", "REQ-012", "LogViewer Effacer", "Bouton 'Effacer' du LogViewer après plusieurs logs", "App démarrée + logs", "—", "Textbox vidée immédiatement, line_count remis à 0", "Mineur"),
    ("T-033", "IHM", "REQ-012", "LogViewer Export", "Bouton 'Export', choisir un .txt", "App démarrée + logs", "—", "Fichier .txt contenant les logs affichés", "Mineur"),
    ("T-034", "IHM", "REQ-012", "LogViewer filtre niveau", "Sélectionner 'WARNING' dans le dropdown niveau", "Logs DEBUG+INFO+WARN+ERROR présents", "—", "Seuls les WARNING et ERROR restent visibles", "Majeur"),
    ("T-035", "IHM", "REQ-012", "LogViewer couleurs", "Provoquer des logs de chaque niveau (start/stop, erreur)", "—", "—", "DEBUG gris, INFO blanc, WARNING ambre, ERROR rouge, SUCCESS vert", "Mineur"),
    ("T-036", "IHM", "REQ-007", "global_progress_label initial 'Prêt'", "Au démarrage de l'app", "—", "—", "Label gauche affiche 'Prêt' en gras", "Mineur"),
    ("T-037", "IHM", "REQ-007", "global_progress_bar visible vide", "Au démarrage", "—", "—", "Barre 20 px haute, bordure visible, vide", "Majeur"),
    ("T-038", "IHM", "REQ-007", "global_progress_percent '0 %' initial", "Au démarrage", "—", "—", "Label droit affiche '0 %'", "Majeur"),
    ("T-039", "IHM", "REQ-014", "Slider FPS - changement de valeur", "Glisser le slider FPS à 60", "Section Vidéo dépliée", "—", "Label valeur affiche '60.0' et la valeur sera utilisée au prochain run", "Mineur"),
    ("T-040", "IHM", "REQ-014", "Dropdown Qualité - changement", "Sélectionner 'Maximum' dans Qualité", "Section Vidéo dépliée", "—", "Valeur stockée, mapping vers preset 'slower' au prochain run (log)", "Majeur"),

    # ============ Paramètres (T-041..T-062) — pairwise + bornes ============
    ("T-041", "Paramètres", "REQ-014", "Pairwise 1/16 : Lineaire/Normal/Basse/Original", "Workflow avec Easing=Lineaire, Fusion=Normal, Qualité=Basse, Résolution=Original", "input_nominal", "input_nominal", "Vidéo produite, preset ffmpeg='ultrafast', easing=LINEAR, blend=ALPHA", "Majeur"),
    ("T-042", "Paramètres", "REQ-014", "Pairwise 2/16 : Lineaire/Cross-dissolve/Moyenne/1080p", "Easing=Lineaire, Fusion=Cross-dissolve, Qualité=Moyenne, Résolution=1080p", "input_nominal", "input_nominal", "Vidéo 1080p, preset='medium', blend=ALPHA (cross-dissolve→alpha)", "Majeur"),
    ("T-043", "Paramètres", "REQ-014", "Pairwise 3/16 : Lineaire/Additive/Haute/720p", "—", "input_nominal", "input_nominal", "Vidéo 720p, preset='slow', blend=ADDITIVE", "Majeur"),
    ("T-044", "Paramètres", "REQ-014", "Pairwise 4/16 : Ease In/Out/Normal/Haute/480p", "—", "input_nominal", "input_nominal", "Vidéo 480p, preset='slow', easing=EASE_IN_OUT", "Majeur"),
    ("T-045", "Paramètres", "REQ-014", "Pairwise 5/16 : Ease In/Out/Cross-dissolve/Maximum/Original", "—", "input_nominal", "input_nominal", "Preset='slower', easing=EASE_IN_OUT", "Majeur"),
    ("T-046", "Paramètres", "REQ-014", "Pairwise 6/16 : Ease In/Out/Additive/Basse/720p", "—", "input_nominal", "input_nominal", "Preset='ultrafast', blend=ADDITIVE", "Majeur"),
    ("T-047", "Paramètres", "REQ-014", "Pairwise 7/16 : Ease In/Normal/Haute/1080p", "—", "input_nominal", "input_nominal", "Easing=EASE_IN, preset='slow', 1080p", "Majeur"),
    ("T-048", "Paramètres", "REQ-014", "Pairwise 8/16 : Ease In/Cross-dissolve/Basse/480p", "—", "input_nominal", "input_nominal", "Easing=EASE_IN, preset='ultrafast', 480p", "Mineur"),
    ("T-049", "Paramètres", "REQ-014", "Pairwise 9/16 : Ease In/Additive/Moyenne/Original", "—", "input_nominal", "input_nominal", "Easing=EASE_IN, blend=ADDITIVE, preset='medium'", "Mineur"),
    ("T-050", "Paramètres", "REQ-014", "Pairwise 10/16 : Ease Out/Normal/Maximum/720p", "—", "input_nominal", "input_nominal", "Easing=EASE_OUT, preset='slower', 720p", "Mineur"),
    ("T-051", "Paramètres", "REQ-014", "Pairwise 11/16 : Ease Out/Cross-dissolve/Haute/Original", "—", "input_nominal", "input_nominal", "Easing=EASE_OUT, preset='slow', original", "Mineur"),
    ("T-052", "Paramètres", "REQ-014", "Pairwise 12/16 : Ease Out/Additive/Maximum/480p", "—", "input_nominal", "input_nominal", "Easing=EASE_OUT, blend=ADDITIVE, preset='slower'", "Mineur"),
    ("T-053", "Paramètres", "REQ-014", "Pairwise 13/16 : Lineaire/Normal/Moyenne/720p", "—", "input_nominal", "input_nominal", "OK", "Mineur"),
    ("T-054", "Paramètres", "REQ-014", "Pairwise 14/16 : Ease In/Out/Normal/Basse/Original", "—", "input_nominal", "input_nominal", "OK", "Mineur"),
    ("T-055", "Paramètres", "REQ-014", "Pairwise 15/16 : Ease In/Cross-dissolve/Maximum/720p", "—", "input_nominal", "input_nominal", "OK", "Mineur"),
    ("T-056", "Paramètres", "REQ-014", "Pairwise 16/16 : Ease Out/Cross-dissolve/Basse/1080p", "—", "input_nominal", "input_nominal", "OK", "Mineur"),
    ("T-057", "Paramètres", "REQ-009", "Slider Tentatives min (1)", "Workflow avec retry_detection=1 sur input difficile", "input avec visage difficile", "—", "Si dlib échoue à la 1ʳᵉ tentative, image marquée 'aucun visage détecté'", "Majeur"),
    ("T-058", "Paramètres", "REQ-009", "Slider Tentatives max (5)", "Workflow avec retry_detection=5", "input_no_face", "input_no_face", "Logs: 'Tentative 1', '2', ..., '5' avant warning final", "Mineur"),
    ("T-059", "Paramètres", "REQ-014", "Slider FPS bornes (10 et 60)", "Deux runs successifs avec FPS=10 puis FPS=60", "input_nominal", "input_nominal", "Vidéo de durée appropriée : à 60 fps, 3 s = 180 frames", "Majeur"),
    ("T-060", "Paramètres", "—", "Slider Transition bornes (0.5 et 10)", "—", "input_nominal", "input_nominal", "Durée vidéo cohérente (transition×fps)", "Mineur"),
    ("T-061", "Paramètres", "—", "Slider Pause bornes (0 et 5)", "—", "input_nominal", "input_nominal", "Pauses bien insérées entre les transitions", "Mineur"),
    ("T-062", "Paramètres", "—", "Slider Bordure bornes (0 et 100)", "—", "input_nominal", "input_nominal", "Bordure blanche visible sur sorties alignées", "Mineur"),

    # ============ Entrées (T-063..T-074) ============
    ("T-063", "Entrées", "REQ-001 / REQ-008", "Workflow nominal synthétique", "Workflow complet avec dossier de visages synthétiques", "Modèle dlib présent", "input_nominal", "Pipeline atteint Morph mais 'aucun visage détecté' (formes géométriques non reconnues par dlib) → fallback cross-dissolve", "Bloquant"),
    ("T-064", "Entrées", "REQ-017", "Dossier source vide", "Sélectionner input_vide puis cliquer Lancer", "—", "input_vide", "Messagebox d'erreur 'Aucune image trouvée dans le dossier source', pas de thread lancé", "Bloquant"),
    ("T-065", "Entrées", "REQ-017", "Dossier avec 1 seule image", "—", "—", "input_1image", "Etape Import OK ; étape Morph raise 'Au moins 2 images nécessaires'", "Majeur"),
    ("T-066", "Entrées", "—", "100 images (volume)", "Workflow nominal sur volume", "—", "input_volume", "Termine sans crash, mémoire < 1 GB, durée raisonnable", "Majeur"),
    ("T-067", "Entrées", "—", "Mauvais format (.png texte)", "—", "—", "input_mauvais_format", "validate_image_file raise CORRUPTED, images skippées avec warning", "Majeur"),
    ("T-068", "Entrées", "REQ-019", "Noms fichiers Unicode", "Workflow nominal", "—", "input_specchars", "Pas de crash, fichiers lus correctement, logs lisibles", "Majeur"),
    ("T-069", "Entrées", "REQ-019", "Path racine avec espaces et #", "Workflow complet depuis D:\\#Bureau\\MorphoLapse\\...", "—", "input_nominal", "OK (déjà validé smoke launch)", "Majeur"),
    ("T-070", "Entrées", "—", "Images 2000×2000 (limite haute)", "—", "—", "input_limite_haute", "Pas de crash, mémoire < 4 GB, vidéo produite à la résolution config", "Majeur"),
    ("T-071", "Entrées", "—", "Images 64×64 (limite basse)", "—", "—", "input_limite_basse", "dlib échoue (image trop petite), fallback cross-dissolve", "Mineur"),
    ("T-072", "Entrées", "—", "PNG corrompus (50 bytes)", "—", "—", "input_corrompu", "validate_image_file raise CORRUPTED, 0 image valide → erreur claire", "Majeur"),
    ("T-073", "Entrées", "REQ-009", "Aucun visage détectable (gradients)", "Workflow nominal", "retry=3", "input_no_face", "Logs 'Aucun visage détecté' pour chaque image, fallback cross-dissolve sur Morph", "Majeur"),
    ("T-074", "Entrées", "REQ-001", "Photos réelles fournies", "Workflow complet avec vraies photos", "Photos déposées par utilisateur", "input_reel/", "Pipeline complet réussit, vidéo de qualité visuelle correcte", "Bloquant"),

    # ============ Sorties (T-075..T-082) ============
    ("T-075", "Sorties", "REQ-010", "Format MP4 H.264 conforme", "Lire la sortie avec ffprobe -show_streams", "Workflow T-074 OK", "—", "codec_name='h264', pix_fmt='yuv420p', container='mp4'", "Bloquant"),
    ("T-076", "Sorties", "REQ-010", "Résolution effective 720p", "Workflow avec Résolution=720p", "input_reel", "—", "Hauteur=720, largeur=arrondi pair selon ratio", "Majeur"),
    ("T-077", "Sorties", "REQ-010", "Résolution 1080p", "—", "—", "—", "Hauteur=1080", "Mineur"),
    ("T-078", "Sorties", "REQ-010", "Résolution 480p", "—", "—", "—", "Hauteur=480", "Mineur"),
    ("T-079", "Sorties", "—", "Export GIF optionnel", "Workflow avec create_gif=True", "FFmpeg dispo", "input_reel", "Fichier 03_morph/morph_preview.gif présent, lisible", "Mineur"),
    ("T-080", "Sorties", "—", "Export thumbnail", "Workflow avec thumbnail=True", "—", "input_reel", "Fichier 03_morph/thumbnail.jpg présent (640 px)", "Mineur"),
    ("T-081", "Sorties", "—", "run_summary.json structure", "Ouvrir runs/<ts>/04_export/run_summary.json", "Workflow OK", "—", "JSON valide, clés : timestamp, run_dir, input, output, config, files", "Majeur"),
    ("T-082", "Sorties", "—", "metadata.txt contenu", "Ouvrir runs/<ts>/04_export/metadata.txt", "Workflow OK", "—", "Texte avec project, version, created, source_images, fps, transition_duration, output_format", "Mineur"),

    # ============ Cas limites (T-083..T-094) ============
    ("T-083", "Cas limites", "REQ-017", "Modèle dlib absent", "Renommer temporairement shape_predictor*.dat, cliquer Lancer", "Sélection input valide", "input_nominal", "Messagebox d'erreur explicite : 'Modèle de détection faciale introuvable'. Aucun thread lancé.", "Bloquant"),
    ("T-084", "Cas limites", "—", "FFmpeg absent du PATH", "Retirer ffmpeg du PATH, démarrer workflow", "Modèle dlib présent", "input_nominal", "Étape Morph raise RuntimeError 'FFmpeg n'est pas disponible'. Workflow s'arrête (ou continue selon option).", "Bloquant"),
    ("T-085", "Cas limites", "REQ-019", "Path output avec Unicode", "Output dir = manual_T085/sortie_éàω_中文", "—", "input_nominal", "Pas de crash, fichiers créés en UTF-8", "Majeur"),
    ("T-086", "Cas limites", "—", "Output dir sans droits écriture", "Output dir = un dossier read-only", "Output dir read-only", "input_nominal", "Erreur d'écriture loguée, messagebox finale 'erreurs'", "Majeur"),
    ("T-087", "Cas limites", "—", "Disque plein simulé", "Remplir un volume puis lancer", "Volume saturé", "input_nominal", "Erreur d'écriture loguée, pas de crash", "Mineur"),
    ("T-088", "Cas limites", "REQ-008", "Étape Import désactivée", "Décocher Import puis Lancer", "—", "input_nominal", "Step Import = SKIPPED ; Align échoue car context.images est vide (à confirmer)", "Mineur"),
    ("T-089", "Cas limites", "REQ-008", "Étape Align désactivée", "Décocher Align", "—", "input_nominal", "Step Align = SKIPPED ; Morph utilise context.images bruts (pas alignés)", "Mineur"),
    ("T-090", "Cas limites", "REQ-008", "Étape Morph désactivée", "Décocher Morph", "—", "input_nominal", "Step Morph = SKIPPED ; Export saute la copie vidéo", "Mineur"),
    ("T-091", "Cas limites", "REQ-008", "Étape Export désactivée", "Décocher Export", "—", "input_nominal", "Step Export = SKIPPED ; vidéo dans runs/<ts>/03_morph/ uniquement", "Mineur"),
    ("T-092", "Cas limites", "—", "Output dir = input dir", "Mettre les deux à la même valeur", "—", "input_nominal", "Pas de crash, copies dans le sous-dossier 04_export", "Mineur"),
    ("T-093", "Cas limites", "REQ-017", "Input dir supprimé entre sélection et Lancer", "Sélectionner puis supprimer le dossier, cliquer Lancer", "—", "input_nominal", "Messagebox 'Le dossier source est introuvable'", "Mineur"),
    ("T-094", "Cas limites", "—", "Reference image inexistante", "Choisir ref puis supprimer le fichier, lancer", "—", "input_nominal", "Step Align utilise context.images[0] en fallback (warning logué)", "Mineur"),

    # ============ Performance (T-095..T-099) ============
    ("T-095", "Performance", "—", "Démarrage GUI < 3 s", "Mesurer le temps entre python main.py et premier mainloop tick", "—", "—", "≤ 3 s en source ; ≤ 2 s en onefile warm ; ≤ 4 s en onefile cold", "Mineur"),
    ("T-096", "Performance", "—", "Réactivité UI pendant run", "Pendant Morph, déplacer la fenêtre, hover boutons", "—", "input_nominal", "UI répond sous 100 ms (callbacks after(0,...))", "Majeur"),
    ("T-097", "Performance", "—", "Mémoire sur 100 images", "Tâche moniteur RAM pendant workflow volume", "—", "input_volume", "Pic mémoire < 1 GB", "Mineur"),
    ("T-098", "Performance", "—", "Durée morph 5 images", "Mesurer durée workflow nominal", "—", "input_nominal", "< 60 s (5 paires × ~10 s)", "Mineur"),
    ("T-099", "Performance", "—", "Durée morph haute résolution", "Workflow sur limite_haute (2000×2000)", "—", "input_limite_haute", "< 5 min, mémoire < 4 GB", "Mineur"),

    # ============ Robustesse (T-100..T-107) ============
    ("T-100", "Robustesse", "REQ-005", "Annulation pendant Import", "Lancer puis Échap immédiat", "—", "input_volume", "Step Import = CANCELLED en < 1 s, étapes suivantes non exécutées", "Bloquant"),
    ("T-101", "Robustesse", "REQ-005", "Annulation pendant Align", "Lancer puis Échap pendant Align", "—", "input_volume", "Step Align = CANCELLED, Import = COMPLETED", "Bloquant"),
    ("T-102", "Robustesse", "REQ-005", "Annulation pendant Morph (loop interne)", "Lancer puis Échap pendant Morph", "—", "input_volume", "Step Morph = CANCELLED, sortie partielle non corrompue", "Majeur"),
    ("T-103", "Robustesse", "REQ-016", "Double-clic Lancer", "Cliquer Lancer 2× en < 100 ms", "—", "input_nominal", "Une seule exécution, garde _workflow_starting active", "Majeur"),
    ("T-104", "Robustesse", "REQ-006", "Ctrl+1 ignoré dans LogViewer", "Cliquer dans la zone de logs, presser Ctrl+1", "—", "—", "Aucun toggle, focus reste dans LogViewer", "Mineur"),
    ("T-105", "Robustesse", "REQ-006", "Ctrl+2 ignoré dans Entry", "Focus dans 'Dossier source', presser Ctrl+2", "—", "—", "Frappe non perturbée, section Morphing non togglée", "Mineur"),
    ("T-106", "Robustesse", "REQ-011", "Reset → Sauver → Restart", "Reset, modifier 1 option, Sauver, fermer, relancer", "—", "—", "Modif unique restaurée, autres options à default", "Mineur"),
    ("T-107", "Robustesse", "—", "Callback Logger qui lève une exception", "Ajouter un callback de test qui raise", "—", "—", "Workflow continue, log warning 'Logger callback raised'", "Mineur"),

    # ============ Régression (T-108..T-118) — bugs corrigés Phase 1 ============
    ("T-108", "Régression", "REQ-006", "T-016 historique Ctrl+1", "Ré-exécution du test originel", "—", "—", "Toggle section Vidéo", "Bloquant"),
    ("T-109", "Régression", "REQ-006", "T-017 historique Ctrl+2", "—", "—", "—", "Toggle section Morphing", "Bloquant"),
    ("T-110", "Régression", "REQ-006", "T-018 historique Ctrl+3", "—", "—", "—", "Toggle section Alignement", "Bloquant"),
    ("T-111", "Régression", "REQ-006", "T-019 historique Ctrl+4", "—", "—", "—", "Toggle section Détection", "Bloquant"),
    ("T-112", "Régression", "REQ-006", "T-020 historique Ctrl+5", "—", "—", "—", "Toggle section Workflow", "Bloquant"),
    ("T-113", "Régression", "REQ-002", "T-030 historique : compteur sur sélection référence", "Sélectionner uniquement la référence", "App vide", "face_000.png", "stats_label se met à jour avec 'Réf: face_000.png'", "Majeur"),
    ("T-114", "Régression", "REQ-003", "T-034 historique : Lancer disabled à vide", "App démarrée sans config", "—", "—", "Lancer disabled", "Bloquant"),
    ("T-115", "Régression", "REQ-007", "T-036 historique : barre visible à 0%", "App démarrée", "—", "—", "Barre visible, bordée, label '0 %'", "Majeur"),
    ("T-116", "Régression", "REQ-005", "T-023 historique : Échap pendant traitement", "Lancer puis Échap", "—", "input_nominal", "Annulation effective < 1 s", "Bloquant"),
    ("T-117", "Régression", "REQ-014", "Mapping FR Basse → preset ultrafast", "Run avec Qualité=Basse, lire les logs", "—", "input_nominal", "Log : 'preset=ultrafast'", "Majeur"),
    ("T-118", "Régression", "REQ-018", "DA-1 : restauration référence après redémarrage", "Choisir ref, Sauver, fermer, relancer", "—", "face_000.png", "Champ référence pré-rempli au démarrage", "Mineur"),
]


HEADERS = [
    "ID_Test",
    "Catégorie",
    "Exigence liée",
    "Fonctionnalité",
    "Description",
    "Pré-requis",
    "Données entrée",
    "Résultat attendu",
    "Résultat obtenu",
    "Statut",
    "Sévérité",
    "Testeur",
    "Date",
    "Commentaires",
]


def _sheet_matrice(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Matrice"

    # En-tête
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # Largeur colonnes
    widths = {
        "A": 9,   # ID
        "B": 12,  # Cat
        "C": 16,  # REQ
        "D": 28,  # Fonctionnalité
        "E": 56,  # Description
        "F": 22,  # Pré-requis
        "G": 22,  # Données entrée
        "H": 56,  # Résultat attendu
        "I": 20,  # Résultat obtenu
        "J": 10,  # Statut
        "K": 11,  # Sévérité
        "L": 12,  # Testeur
        "M": 12,  # Date
        "N": 30,  # Commentaires
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Lignes
    for row_idx, t in enumerate(TESTS, start=2):
        (test_id, cat, req, func, desc, pre, data_in, expected, severity) = t
        ws.cell(row=row_idx, column=1, value=test_id).alignment = CENTER
        ws.cell(row=row_idx, column=2, value=cat).alignment = CENTER
        ws.cell(row=row_idx, column=3, value=req).alignment = CENTER
        ws.cell(row=row_idx, column=4, value=func).alignment = LEFT_WRAP
        ws.cell(row=row_idx, column=5, value=desc).alignment = LEFT_WRAP
        ws.cell(row=row_idx, column=6, value=pre).alignment = LEFT_WRAP
        ws.cell(row=row_idx, column=7, value=data_in).alignment = LEFT_WRAP
        ws.cell(row=row_idx, column=8, value=expected).alignment = LEFT_WRAP
        # I = Résultat obtenu (vide, rempli par testeur)
        # J = Statut (vide → choisir parmi OK / NOK / NA)
        ws.cell(row=row_idx, column=11, value=severity).alignment = CENTER
        # L = Testeur, M = Date, N = Commentaires (vide)

        # Couleur de fond de la cellule "Catégorie"
        fill = CAT_FILL.get(cat)
        if fill:
            ws.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor=fill)

    # Freeze la première ligne et la colonne ID
    ws.freeze_panes = "B2"


def _sheet_synthese(wb: Workbook) -> None:
    ws = wb.create_sheet("Synthèse")

    ws["A1"] = "MorphoLapse v2.0.0 — Synthèse qualification"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Généré le {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="FF666666")

    # Tableau récap par catégorie
    cats = list(CAT_FILL.keys())
    headers = ["Catégorie", "Total", "OK", "NOK", "NA", "À tester", "% OK"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    for i, cat in enumerate(cats, start=5):
        ws.cell(row=i, column=1, value=cat).alignment = CENTER
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=CAT_FILL[cat])
        # Total = compte de la catégorie dans la feuille Matrice
        ws.cell(row=i, column=2, value=f'=COUNTIF(Matrice!B:B,"{cat}")')
        ws.cell(row=i, column=3, value=f'=COUNTIFS(Matrice!B:B,"{cat}",Matrice!J:J,"OK")')
        ws.cell(row=i, column=4, value=f'=COUNTIFS(Matrice!B:B,"{cat}",Matrice!J:J,"NOK")')
        ws.cell(row=i, column=5, value=f'=COUNTIFS(Matrice!B:B,"{cat}",Matrice!J:J,"NA")')
        ws.cell(row=i, column=6, value=f"=B{i}-C{i}-D{i}-E{i}")
        ws.cell(row=i, column=7, value=f'=IF(B{i}=0,"",ROUND(C{i}/B{i}*100,1))')
        ws.cell(row=i, column=7).number_format = "0.0"

    # Ligne TOTAL
    total_row = 5 + len(cats)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 7):
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({letter}5:{letter}{total_row - 1})").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=f"=IF(B{total_row}=0,\"\",ROUND(C{total_row}/B{total_row}*100,1))").font = Font(bold=True)

    # Tableau par sévérité
    ws["A15"] = "Par sévérité"
    ws["A15"].font = Font(bold=True, size=12)
    sev_headers = ["Sévérité", "Total", "OK", "NOK", "% OK"]
    for col_idx, h in enumerate(sev_headers, start=1):
        c = ws.cell(row=16, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    severities = ["Bloquant", "Majeur", "Mineur"]
    for i, sev in enumerate(severities, start=17):
        ws.cell(row=i, column=1, value=sev).alignment = CENTER
        ws.cell(row=i, column=2, value=f'=COUNTIF(Matrice!K:K,"{sev}")')
        ws.cell(row=i, column=3, value=f'=COUNTIFS(Matrice!K:K,"{sev}",Matrice!J:J,"OK")')
        ws.cell(row=i, column=4, value=f'=COUNTIFS(Matrice!K:K,"{sev}",Matrice!J:J,"NOK")')
        ws.cell(row=i, column=5, value=f'=IF(B{i}=0,"",ROUND(C{i}/B{i}*100,1))')
        ws.cell(row=i, column=5).number_format = "0.0"

    # Largeur colonnes
    for col, w in {"A": 22, "B": 10, "C": 10, "D": 10, "E": 10, "F": 12, "G": 10}.items():
        ws.column_dimensions[col].width = w

    # Note
    ws["A23"] = "Légende Statut : OK / NOK / NA / (vide = à tester)"
    ws["A23"].font = Font(italic=True, color="FF666666")


def main() -> int:
    wb = Workbook()
    _sheet_matrice(wb)
    _sheet_synthese(wb)
    wb.save(OUT)
    print(f"OK : {OUT}  ({len(TESTS)} tests)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
