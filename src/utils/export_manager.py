"""
Export Manager - Module d'export universel
Supporte: Excel (.xlsx), PDF, JSON, CSV, ZIP
"""

import os
import json
import csv
import zipfile
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import List, Dict, Any, Optional, Union
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


@dataclass
class ExportResult:
    """Résultat d'un export"""
    success: bool
    file_path: str
    format: str
    size_bytes: int
    message: str
    timestamp: str = ""

    def __post_init__(self):
        if not self.timestamp:
            self.timestamp = datetime.now().isoformat()


@dataclass
class ExportOptions:
    """Options d'export configurables"""
    # Général
    output_dir: str = ""
    filename_prefix: str = "export"
    include_timestamp: bool = True

    # Excel
    excel_sheet_name: str = "Données"
    excel_freeze_header: bool = True
    excel_auto_width: bool = True
    excel_header_color: str = "366092"
    excel_alternate_rows: bool = True

    # PDF
    pdf_page_size: str = "A4"
    pdf_title: str = "Rapport"
    pdf_author: str = "MorphoLapse"
    pdf_include_footer: bool = True

    # JSON
    json_indent: int = 2
    json_ensure_ascii: bool = False

    # CSV
    csv_delimiter: str = ";"
    csv_encoding: str = "utf-8-sig"  # BOM pour Excel
    csv_quoting: int = csv.QUOTE_MINIMAL


class ExportManager:
    """
    Gestionnaire d'export universel pour MorphoLapse

    Supporte:
    - Excel (.xlsx) avec mise en forme professionnelle
    - PDF avec tableaux stylisés
    - JSON structuré
    - CSV compatible Excel
    - Archives ZIP
    """

    def __init__(self, options: Optional[ExportOptions] = None):
        self.options = options or ExportOptions()
        self._last_export: Optional[ExportResult] = None

    @property
    def excel_available(self) -> bool:
        return OPENPYXL_AVAILABLE

    @property
    def pdf_available(self) -> bool:
        return REPORTLAB_AVAILABLE

    def _generate_filename(self, extension: str) -> str:
        """Génère un nom de fichier avec timestamp optionnel"""
        base = self.options.filename_prefix
        if self.options.include_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base = f"{base}_{timestamp}"
        return f"{base}.{extension}"

    def _ensure_output_dir(self) -> str:
        """Assure que le dossier de sortie existe"""
        output_dir = self.options.output_dir or os.getcwd()
        os.makedirs(output_dir, exist_ok=True)
        return output_dir

    # ========== EXPORT EXCEL ==========

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: Optional[str] = None
    ) -> ExportResult:
        """
        Exporte des données vers un fichier Excel formaté professionnellement

        Args:
            data: Liste de dictionnaires (chaque dict = une ligne)
            filename: Nom du fichier (auto-généré si non fourni)
            sheet_name: Nom de la feuille

        Returns:
            ExportResult avec statut et chemin
        """
        if not OPENPYXL_AVAILABLE:
            return ExportResult(
                success=False,
                file_path="",
                format="xlsx",
                size_bytes=0,
                message="Module openpyxl non installé. Installez-le avec: pip install openpyxl"
            )

        if not data:
            return ExportResult(
                success=False,
                file_path="",
                format="xlsx",
                size_bytes=0,
                message="Aucune donnée à exporter"
            )

        try:
            output_dir = self._ensure_output_dir()
            filename = filename or self._generate_filename("xlsx")
            filepath = os.path.join(output_dir, filename)
            sheet_name = sheet_name or self.options.excel_sheet_name

            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]  # Limite Excel

            # En-têtes
            headers = list(data[0].keys())

            # Style des en-têtes
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color=self.options.excel_header_color,
                end_color=self.options.excel_header_color,
                fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Écrire les en-têtes
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Couleurs alternées pour les lignes
            alt_fill = PatternFill(
                start_color="F2F2F2",
                end_color="F2F2F2",
                fill_type="solid"
            )

            # Écrire les données
            for row_idx, row_data in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col, value=row_data.get(header, ""))
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

                    if self.options.excel_alternate_rows and row_idx % 2 == 0:
                        cell.fill = alt_fill

            # Ajuster la largeur des colonnes
            if self.options.excel_auto_width:
                for col_idx, header in enumerate(headers, 1):
                    max_length = len(str(header))
                    for row in range(2, len(data) + 2):
                        cell_value = ws.cell(row=row, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # Figer la première ligne
            if self.options.excel_freeze_header:
                ws.freeze_panes = "A2"

            # Sauvegarder
            wb.save(filepath)

            file_size = os.path.getsize(filepath)
            self._last_export = ExportResult(
                success=True,
                file_path=filepath,
                format="xlsx",
                size_bytes=file_size,
                message=f"Export Excel réussi: {len(data)} lignes"
            )
            return self._last_export

        except Exception as e:
            return ExportResult(
                success=False,
                file_path="",
                format="xlsx",
                size_bytes=0,
                message=f"Erreur export Excel: {str(e)}"
            )

    # ========== EXPORT PDF ==========

    def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        title: Optional[str] = None
    ) -> ExportResult:
        """
        Exporte des données vers un fichier PDF formaté

        Args:
            data: Liste de dictionnaires
            filename: Nom du fichier
            title: Titre du document

        Returns:
            ExportResult
        """
        if not REPORTLAB_AVAILABLE:
            return ExportResult(
                success=False,
                file_path="",
                format="pdf",
                size_bytes=0,
                message="Module reportlab non installé. Installez-le avec: pip install reportlab"
            )

        if not data:
            return ExportResult(
                success=False,
                file_path="",
                format="pdf",
                size_bytes=0,
                message="Aucune donnée à exporter"
            )

        try:
            output_dir = self._ensure_output_dir()
            filename = filename or self._generate_filename("pdf")
            filepath = os.path.join(output_dir, filename)
            title = title or self.options.pdf_title

            # Page size
            page_size = A4 if self.options.pdf_page_size == "A4" else letter

            # Créer le document
            doc = SimpleDocTemplate(
                filepath,
                pagesize=page_size,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )

            elements = []
            styles = getSampleStyleSheet()

            # Titre
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=20,
                alignment=1  # Center
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.5*cm))

            # Sous-titre avec date
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                alignment=1
            )
            date_str = datetime.now().strftime("%d/%m/%Y à %H:%M")
            elements.append(Paragraph(f"Généré le {date_str}", subtitle_style))
            elements.append(Spacer(1, 1*cm))

            # Préparer les données du tableau
            headers = list(data[0].keys())
            table_data = [headers]

            for row in data:
                table_data.append([str(row.get(h, "")) for h in headers])

            # Créer le tableau
            table = Table(table_data, repeatRows=1)

            # Style du tableau
            style = TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),

                # Corps
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),

                # Grille
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

                # Alternance couleurs
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
            ])
            table.setStyle(style)
            elements.append(table)

            # Footer
            if self.options.pdf_include_footer:
                elements.append(Spacer(1, 1*cm))
                footer_style = ParagraphStyle(
                    'Footer',
                    parent=styles['Normal'],
                    fontSize=8,
                    textColor=colors.grey,
                    alignment=1
                )
                elements.append(Paragraph(
                    f"Généré par {self.options.pdf_author} | {len(data)} enregistrements",
                    footer_style
                ))

            # Générer le PDF
            doc.build(elements)

            file_size = os.path.getsize(filepath)
            self._last_export = ExportResult(
                success=True,
                file_path=filepath,
                format="pdf",
                size_bytes=file_size,
                message=f"Export PDF réussi: {len(data)} lignes"
            )
            return self._last_export

        except Exception as e:
            return ExportResult(
                success=False,
                file_path="",
                format="pdf",
                size_bytes=0,
                message=f"Erreur export PDF: {str(e)}"
            )

    # ========== EXPORT JSON ==========

    def export_to_json(
        self,
        data: Union[List[Dict], Dict],
        filename: Optional[str] = None,
        pretty: bool = True
    ) -> ExportResult:
        """
        Exporte des données vers un fichier JSON

        Args:
            data: Données à exporter
            filename: Nom du fichier
            pretty: Formatage indenté

        Returns:
            ExportResult
        """
        try:
            output_dir = self._ensure_output_dir()
            filename = filename or self._generate_filename("json")
            filepath = os.path.join(output_dir, filename)

            indent = self.options.json_indent if pretty else None

            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(
                    data, f,
                    indent=indent,
                    ensure_ascii=self.options.json_ensure_ascii,
                    default=str  # Pour datetime et autres
                )

            file_size = os.path.getsize(filepath)
            count = len(data) if isinstance(data, list) else 1

            self._last_export = ExportResult(
                success=True,
                file_path=filepath,
                format="json",
                size_bytes=file_size,
                message=f"Export JSON réussi: {count} éléments"
            )
            return self._last_export

        except Exception as e:
            return ExportResult(
                success=False,
                file_path="",
                format="json",
                size_bytes=0,
                message=f"Erreur export JSON: {str(e)}"
            )

    # ========== EXPORT CSV ==========

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> ExportResult:
        """
        Exporte des données vers un fichier CSV compatible Excel

        Args:
            data: Liste de dictionnaires
            filename: Nom du fichier

        Returns:
            ExportResult
        """
        if not data:
            return ExportResult(
                success=False,
                file_path="",
                format="csv",
                size_bytes=0,
                message="Aucune donnée à exporter"
            )

        try:
            output_dir = self._ensure_output_dir()
            filename = filename or self._generate_filename("csv")
            filepath = os.path.join(output_dir, filename)

            headers = list(data[0].keys())

            with open(filepath, 'w', newline='', encoding=self.options.csv_encoding) as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=headers,
                    delimiter=self.options.csv_delimiter,
                    quoting=self.options.csv_quoting
                )
                writer.writeheader()
                writer.writerows(data)

            file_size = os.path.getsize(filepath)
            self._last_export = ExportResult(
                success=True,
                file_path=filepath,
                format="csv",
                size_bytes=file_size,
                message=f"Export CSV réussi: {len(data)} lignes"
            )
            return self._last_export

        except Exception as e:
            return ExportResult(
                success=False,
                file_path="",
                format="csv",
                size_bytes=0,
                message=f"Erreur export CSV: {str(e)}"
            )

    # ========== EXPORT ZIP ==========

    def create_archive(
        self,
        files: List[str],
        filename: Optional[str] = None,
        compression: int = zipfile.ZIP_DEFLATED
    ) -> ExportResult:
        """
        Crée une archive ZIP de fichiers

        Args:
            files: Liste des chemins de fichiers à archiver
            filename: Nom de l'archive
            compression: Niveau de compression

        Returns:
            ExportResult
        """
        if not files:
            return ExportResult(
                success=False,
                file_path="",
                format="zip",
                size_bytes=0,
                message="Aucun fichier à archiver"
            )

        try:
            output_dir = self._ensure_output_dir()
            filename = filename or self._generate_filename("zip")
            filepath = os.path.join(output_dir, filename)

            added_count = 0
            with zipfile.ZipFile(filepath, 'w', compression) as zf:
                for file_path in files:
                    if os.path.exists(file_path):
                        arcname = os.path.basename(file_path)
                        zf.write(file_path, arcname)
                        added_count += 1

            file_size = os.path.getsize(filepath)
            self._last_export = ExportResult(
                success=True,
                file_path=filepath,
                format="zip",
                size_bytes=file_size,
                message=f"Archive créée: {added_count} fichiers"
            )
            return self._last_export

        except Exception as e:
            return ExportResult(
                success=False,
                file_path="",
                format="zip",
                size_bytes=0,
                message=f"Erreur création archive: {str(e)}"
            )

    # ========== EXPORT RAPPORT WORKFLOW ==========

    def export_workflow_report(
        self,
        workflow_data: Dict[str, Any],
        format: str = "xlsx"
    ) -> ExportResult:
        """
        Exporte un rapport complet du workflow MorphoLapse

        Args:
            workflow_data: Données du workflow (contexte, résultats, logs)
            format: Format de sortie (xlsx, pdf, json)

        Returns:
            ExportResult
        """
        # Préparer les données
        report_data = []

        # Informations générales
        if 'context' in workflow_data:
            ctx = workflow_data['context']
            report_data.append({
                'Section': 'Configuration',
                'Paramètre': 'Dossier source',
                'Valeur': ctx.get('input_dir', 'N/A'),
                'Status': '✓'
            })
            report_data.append({
                'Section': 'Configuration',
                'Paramètre': 'Dossier sortie',
                'Valeur': ctx.get('output_dir', 'N/A'),
                'Status': '✓'
            })
            report_data.append({
                'Section': 'Configuration',
                'Paramètre': 'FPS',
                'Valeur': str(ctx.get('config', {}).get('fps', 25)),
                'Status': '✓'
            })

        # Statistiques des étapes
        if 'steps' in workflow_data:
            for step in workflow_data['steps']:
                status = '✓' if step.get('status') == 'completed' else '✗'
                report_data.append({
                    'Section': 'Étapes',
                    'Paramètre': step.get('name', 'Unknown'),
                    'Valeur': step.get('duration', 'N/A'),
                    'Status': status
                })

        # Images traitées
        if 'images' in workflow_data:
            for img in workflow_data['images']:
                report_data.append({
                    'Section': 'Images',
                    'Paramètre': img.get('filename', 'Unknown'),
                    'Valeur': f"{img.get('landmarks_count', 0)} landmarks",
                    'Status': '✓' if img.get('processed') else '✗'
                })

        # Exporter selon le format
        self.options.filename_prefix = "workflow_report"

        if format == "xlsx":
            return self.export_to_excel(report_data, sheet_name="Rapport Workflow")
        elif format == "pdf":
            return self.export_to_pdf(report_data, title="Rapport de Workflow MorphoLapse")
        elif format == "json":
            return self.export_to_json(workflow_data)
        else:
            return self.export_to_csv(report_data)

    def get_last_export(self) -> Optional[ExportResult]:
        """Retourne le dernier export effectué"""
        return self._last_export

    def get_available_formats(self) -> List[str]:
        """Retourne la liste des formats disponibles"""
        formats = ["csv", "json", "zip"]
        if OPENPYXL_AVAILABLE:
            formats.insert(0, "xlsx")
        if REPORTLAB_AVAILABLE:
            formats.insert(1, "pdf")
        return formats
